import datetime
import io
import json
import os
import time
import zipfile
from urllib.parse import quote as urlquote

import openpyxl as openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import (HttpResponseRedirect, get_object_or_404,
                              redirect, render)
from django.urls import reverse
from selenium import webdriver

from .forms import CadastralNumberForm, OrderFileForm, OrderForm
from .map_funcs import add_table, generate_docx, get_map, get_map_screenshot
from .models import (Area, City, Order, OrderFile, PurposeBuilding, Region,
                     get_image_path)
from .rosreestr2 import GetArea
from .validators import validate_number
from pypdf import PdfReader 
from io import StringIO

# driver = webdriver.Chrome()
User = get_user_model()


def region_autocomplete(request):
    if 'term' in request.GET:
        qs = Region.objects.filter(name__icontains=request.GET.get('term'))
        regions = []
        for region in qs:
            regions.append(region.name)

        return JsonResponse(regions, safe=False)


def area_autocomplete(request):
    qs = Region.objects.get(name__icontains=request.GET.get('region'))
    areas = []
    for area in qs.areas.all():
        areas.append(f'{qs.name}, {area.name}')

    return JsonResponse(areas, safe=False)


def city_autocomplete(request):
    data = request.GET.get('region').split(', ')
    data.remove('')
    region, area = data
    qs = Area.objects.get(name=area)
    citys = []
    for city in qs.citys.all():
        citys.append(f'{region}, {area}, {city.name}')

    return JsonResponse(citys, safe=False)


def purpose_building_autocomplete(request):
    purpose_buildings = PurposeBuilding.objects.all().values_list('purpose', flat=True)
    purpose_building_options = list(purpose_buildings)

    return JsonResponse(purpose_building_options, safe=False)


def ajax_validate_cadastral_number(request):
    cadastral_number = request.GET.get('cadastral_number', None)
    try:
        validate_number(cadastral_number)
        response = {
            'is_valid': True
        }
    except ValidationError:
        response = {
            'is_valid': False
        }

    return JsonResponse(response)


def view_order_cadastral(request, company_slug: str, company_number_slug: str):
    context = {}
    response = HttpResponseRedirect(
        reverse('zakaz:order',
                args=[company_slug, company_number_slug]))

    request.session.modified = True
    try:
        request.session.pop('cadastral_numbers')
        request.session.pop('address')
    except:
        pass

    if 'cadastral_numbers' in request.POST:
        cadastral_numbers = request.POST.getlist('cadastral_numbers')
        request.session['cadastral_numbers'] = cadastral_numbers

        return response

    if 'address' in request.POST:
        address = request.POST.getlist('address')
        request.session['address'] = address

        return response
    
    if 'files' in request.FILES:
        files = request.FILES.getlist('files')
        cadastral_numbers = []
        for file in files:
            file_extension = os.path.splitext(file.name)[-1]
            if file_extension == '.pdf':
                reader = PdfReader(file)
                page = reader.pages[0]
                pdf_text = StringIO(page.extract_text())
                for elem in pdf_text:
                    if 'Кадастровый номер' in elem.strip():
                        cadastral = elem.strip().split(' ')[-1]
                        cadastral_numbers.append(cadastral)
                request.session['cadastral_numbers'] = cadastral_numbers          
            else:
                file = request.FILES.get('files').close()
                messages.error(request, 'Ошибка обработки файла')
                break

        if cadastral_numbers:
            return response

        else:
            form = CadastralNumberForm()

    else:
        form = CadastralNumberForm()

    context['form'] = form

    return render(request, 'zakaz/customer_home.html', context=context)


def view_order(request, company_slug: str, company_number_slug: str):
    coordinates = []
    context = {}
    square_cadastral_area = []

    user_company = get_object_or_404(
        User, company_number_slug=company_number_slug
    )
    cadastral_numbers = request.session['cadastral_numbers'] if 'cadastral_numbers' in request.session else None
    address = request.session['address'] if 'address' in request.session else None

    if address:
        region, area, city = address[0].split(', ')

    if cadastral_numbers:
        cadastral_region = Region.objects.get(
            cadastral_region_number=cadastral_numbers[0].split(':')[0])
        cadastral_area = Area.objects.get(
            cadastral_area_number=cadastral_numbers[0].split(':')[1])

        for number in cadastral_numbers:
            try:
                areas = GetArea(number)
                square_cadastral_area.append(areas.attrs['area_value'])
                coordinates += areas.get_coord()
            except:
                pass

    if request.method == 'POST':
        order_form = OrderForm(request.POST)
        order_files_form = OrderFileForm(request.POST, request.FILES)
        if order_form.is_valid() and order_files_form.is_valid():
            order = order_form.save()
            order.user = user_company
            if cadastral_numbers:
                order.coordinates = coordinates
                order.cadastral_numbers = cadastral_numbers

                tmp_html = os.path.join(
                    settings.BASE_DIR, 'tmp', f'map-{order.id}.html')
                tmp_png = os.path.join(
                    settings.BASE_DIR, 'tmp', f'map-{order.id}.png')

                get_map_screenshot(order.cadastral_numbers).save(tmp_html)

                driver = webdriver.Chrome()
                driver.get(f'file://{tmp_html}')
                time.sleep(1)
                driver.save_screenshot(tmp_png)
                driver.quit()

                img_path = get_image_path(order, 'map.png')
                with open(tmp_png, 'rb') as f:
                    order.map.save(img_path, ContentFile(f.read()), save=True)

                os.remove(tmp_html)
                os.remove(tmp_png)

            order.save()

            for file in request.FILES.getlist('file'):
                OrderFile.objects.create(order=order, file=file)
            messages.success(request, 'Ваша заявка отправлена')

            return HttpResponseRedirect(reverse('zakaz:cadastral', args=[company_slug, company_number_slug]))

        else:
            pass
            # messages.error(request, 'Проверьте правильность введённых данных')
    else:
        order_form = OrderForm(initial={
            'cadastral_numbers': cadastral_numbers if cadastral_numbers else None,
            'region': cadastral_region.id if cadastral_numbers else Region.objects.get(name=region).id,
            'area': cadastral_area.id if cadastral_numbers else Area.objects.get(name=area).id,
            'city': None if cadastral_numbers else City.objects.get(name=city).id,
            'square_unit': Order.SQUARE_UNIT[0][0],
        })

        order_files_form = OrderFileForm()

    context['squares'] = square_cadastral_area
    context['user_company'] = user_company
    context['order_form'] = order_form
    context['order_files_form'] = order_files_form
    context['purpose_building'] = PurposeBuilding.objects.all()
    context['cadastral_numbers'] = cadastral_numbers

    return render(request, 'zakaz/order.html', context=context)


@user_passes_test(lambda u: u.is_staff, login_url='users:company_login')
def view_order_pages(request, company_number_slug: str):
    orders = Order.objects.filter(
        user__company_number_slug=company_number_slug
    ).select_related(
        'city', 'area', 'region', 'work_objective', 'user'
    )
    context = {
        "orders": orders,
    }

    return render(request, 'zakaz/order_pages.html', context=context)


@user_passes_test(lambda u: u.is_staff, login_url='users:company_login')
def view_change_order_status(request, order_id: int):
    square_cadastral_area = []
    order = get_object_or_404(Order.objects.select_related(
        'city', 'area', 'region', 'work_objective', 'user'),
        id=order_id)
    files = OrderFile.objects.select_related('order').filter(order=order.pk)
    if order.cadastral_numbers:
        map_html = get_map(order.cadastral_numbers)
    else:
        map_html = False

    # document_cipher = f"{datetime.datetime.now().strftime('%Y%m%d')}-{order.pk:03d}"
    # screenshot_name = f"{document_cipher}-map"
    # document_igi_name_upload = f'{document_cipher}-igi'
    # document_igdi_name_upload = f'{document_cipher}-igdi'

    if request.method == 'POST':
        order_form = OrderForm(request.POST, instance=order)
        if order_form.is_valid():
            order.object_name = request.POST.get(
                'object_name'
            )
            new_cadastral = request.POST.getlist(
                'new_cadastral_numbers'
            )

            if new_cadastral[0]:
                order.cadastral_numbers += new_cadastral
            else:
                order.cadastral_numbers = request.POST.getlist(
                    'cadastral_numbers'
                )

            for i in order.cadastral_numbers:
                areas = GetArea(i)
                square_cadastral_area.append(areas.attrs['area_value'])
            if request.POST.get('square_unit') == "hectometer":
                order.square = sum(square_cadastral_area) / 1000
            else:
                order.square = sum(square_cadastral_area)

            order = order_form.save()
            order.user.company_number_slug
            return JsonResponse({'success': True})
    else:
        order_form = OrderForm(instance=order)

    context = {
        'type_works': order.type_work.all(),
        'files': files,
        'order_form': order_form,
        'order': order,
        'map_html': map_html,
        'lengt_unit': order.get_length_unit_display(),
        # 'screenshot_name': screenshot_name,
        # 'document_igi_name_upload': document_igi_name_upload,
        # 'document_igdi_name_upload': document_igdi_name_upload
    }

    return render(request, 'zakaz/change_order_status.html', context=context)


def view_rates(request):
    refer = request.META.get('HTTP_REFERER')
    return render(request, 'zakaz/rates.html', context={'refer': refer})


# Скачивание DOCX
def download_docx(request, document_name: str, document_path: str, document_cipher: str, placeholders: dict,
                  coordinates_dict: dict):
    document = generate_docx(document_path, placeholders)

    add_table(document, coordinates_dict)

    output = io.BytesIO()
    document.save(output)
    output.seek(0)

    document_name_upload = f'{document_cipher}-{document_name}'

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{document_name_upload}.docx"'
    return response


# Скачиваем ШИФР-ИГИ
def download_igi_docx(request, pk: int):
    order = get_object_or_404(Order, pk=pk)
    department = order.region.region_department.first()
    location = f"{order.region}, {order.area}, {order.city}, {order.street}, д.{order.house_number}"
    if order.building:
        location += f" {order.building}"

    date = datetime.datetime.now()
    document_cipher = f"{date.strftime('%Y%m%d')}-{order.pk:03d}"

    cadastral_numbers = order.cadastral_numbers
    coordinates = json.loads(order.coordinates)

    coordinates_dict = {}

    for i, coords in enumerate(coordinates):
        if i < len(cadastral_numbers):
            cadastral_num = cadastral_numbers[i]
            coordinates_dict[cadastral_num] = coords[0]

    document_name = 'IGI'
    document_path = os.path.join(settings.MEDIA_ROOT, f'{document_name}.docx')

    if department:
        placeholders = {
            '_шифр-иги': f'{document_cipher}-ИГИ',
            '_должность_руководителя_ведомства': department.director_position,
            '_название_ведомства': department.name,
            '_фио_руководителя_ведомства': f'{department.director_surname} {department.director_name} {department.director_patronymic}',
            '_тел_ведомства': str(department.phone_number),
            '_почта_ведомства': department.email,
            '_дата_текущая': date.strftime("%Y-%m-%d"),
            '_имя_руководителя_ведомства': department.director_name,
            '_название_объекта_полное': order.object_name,
            '_местоположение_объекта': location,
            '_кадастровый_номер': ', '.join(cadastral_numbers),
            '_шифр-тема': f'{document_cipher}-ИГИ'
        }
        if order.map:
            placeholders['_обзорная_схема'] = order.map.path
    else:
        placeholders = {}

    return download_docx(request, document_name, document_path, document_cipher, placeholders, coordinates_dict)


# Скачиваем ШИФР-ИГДИ
def download_igdi_docx(request, pk: int):
    order = get_object_or_404(Order, pk=pk)
    department = order.region.region_department.first()
    location = f"{order.region}, {order.area}, {order.city}, {order.street}, д.{order.house_number}"
    if order.building:
        location += f" {order.building}"

    date = datetime.datetime.now()
    document_cipher = f"{date.strftime('%Y%m%d')}-{order.pk:03d}"

    cadastral_numbers = order.cadastral_numbers
    coordinates = json.loads(order.coordinates)

    coordinates_dict = {}

    for i, coords in enumerate(coordinates):
        if i < len(cadastral_numbers):
            cadastral_num = cadastral_numbers[i]
            coordinates_dict[cadastral_num] = coords[0]

    document_name = 'IGDI'
    document_path = os.path.join(settings.MEDIA_ROOT, f'{document_name}.docx')

    if department:
        placeholders = {
            '_шифр-игди': f'{document_cipher}-ИГДИ',
            '_должность_руководителя_ведомства': department.director_position,
            '_название_ведомства': department.name,
            '_фио_руководителя_ведомства': f'{department.director_surname} {department.director_name} {department.director_patronymic}',
            '_тел_ведомства': str(department.phone_number),
            '_почта_ведомства': department.email,
            '_дата_текущая': date.strftime("%Y-%m-%d"),
            '_имя_руководителя_ведомства': department.director_name,
            '_название_объекта_полное': order.object_name,
            '_местоположение_объекта': location,
            '_кадастровый_номер': ', '.join(cadastral_numbers),
            '_шифр-тема': f'{document_cipher}-ИГДИ'
        }
        if order.map:
            placeholders['_обзорная_схема'] = order.map.path
    else:
        placeholders = {}

    return download_docx(request, document_name, document_path, document_cipher, placeholders, coordinates_dict)


# Скачиваем архив с документами
def download_all_docx(request, pk: int):
    igi_docx = download_igi_docx(request, pk)
    igdi_docx = download_igdi_docx(request, pk)

    document_cipher = f"{datetime.datetime.now().strftime('%Y%m%d')}-{pk:03d}"

    # Создаем архив и добавляем файлы в него
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w") as archive:
        archive.writestr(f"{document_cipher}-igi.docx", igi_docx.content)
        archive.writestr(f"{document_cipher}-igdi.docx", igdi_docx.content)

    # Возвращаем архив пользователю
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = f"attachment; filename={document_cipher}.zip"
    return response


# Скачиваем карты
def download_map(request, pk: int):
    order = get_object_or_404(Order, pk=pk)
    try:
        file_path = order.map.path
        with open(file_path, 'rb') as fh:
            response = HttpResponse(
                fh.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename=' + \
                                              os.path.basename(file_path)
            return response
    except:
        raise Http404


# Скачиваем excel файл
def download_xlsx(request, pk: int):
    order = get_object_or_404(Order, pk=pk)
    cadastral_numbers = order.cadastral_numbers

    coordinates_list = json.loads(order.coordinates)
    coordinates_dict = {}

    # Добавляем проверку на длину списка cadastral_numbers
    if len(cadastral_numbers) == 0:
        return HttpResponse('Нет кадастровых номеров для данного заказа')

    for i, coords in enumerate(coordinates_list):
        # Добавляем проверку на длину списка cadastral_numbers
        if i >= len(cadastral_numbers):
            break

        cadastral_num = cadastral_numbers[i]
        coordinates_dict[cadastral_num] = coords[0]

    # Создаем новый файл Excel
    workbook = openpyxl.Workbook()

    # Получаем лист по умолчанию
    sheet = workbook.active

    # Записываем заголовки строк
    current_row = 1

    # Проходим по каждому ключу словаря coordinates_dict
    for key, coords in coordinates_dict.items():
        # Записываем заголовок строки с координатами углов участка
        sheet.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=3)
        sheet.cell(row=current_row, column=1,
                   value=f'Координаты углов участка {key}')
        current_row += 1

        # Записываем заголовки столбцов таблицы
        sheet.cell(row=current_row, column=1, value='Номер точки')
        sheet.cell(row=current_row, column=2, value='Координата Х')
        sheet.cell(row=current_row, column=3, value='Координата У')
        current_row += 1

        # Проходим по каждой координате в списке для данного ключа
        for i, coord in enumerate(coords):
            # Записываем номер точки, координату Х и координату У
            sheet.cell(row=current_row, column=1, value=str(i + 1))
            sheet.cell(row=current_row, column=2, value=str(coord[0]))
            sheet.cell(row=current_row, column=3, value=str(coord[1]))
            current_row += 1

        # Добавляем пустую строку между таблицами
        current_row += 1

    document_cipher = f"{datetime.datetime.now().strftime('%Y%m%d')}-{order.pk:03d}"

    # Сохраняем файл и отправляем его пользователю
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=coord.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{document_cipher}-coord.xlsx"'
    workbook.save(response)

    return response


def get_message(request):
    return messages.success(request, 'ok')
